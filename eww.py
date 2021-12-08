import os
import cv2
import datetime
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, Color, Alignment, Border, Side, colors
from pyzbar.pyzbar import decode

bold_font = Font(bold=True )
not_bold_font = Font(bold=False )
big_red_text = Font(color='00FF0000')
big_green_text = Font(color='0000FF00')
webcam = True

workbook = load_workbook(filename = 'sample.xlsx')
sheet = workbook.active

img = cv2.imread('2.jpg')
cap = cv2.VideoCapture(0)
cap.set(3,640)
cap.set(4,480)

with open('mydata.txt') as f:
    mydatalist = f.read().splitlines()
sheet.cell(row=1,column=1).value = 'Scholar no'
sheet.cell(row=1,column=1).font = bold_font
var = 2

for ele in mydatalist:
    sheet.cell(row=var,column=1).value = ele
    sheet.cell(row=var,column=1).font = not_bold_font
    var = var + 1

while True:
    if webcam:
        success , img = cap.read()
    for barcode in decode(img):
        mydata = barcode.data.decode('UTF-8')
        print(mydata)

        x = datetime.datetime.now()

        if x.strftime('%x') != sheet.cell(row=1,column=2).value :
            sheet.insert_cols(idx= 2)
            sheet.cell(row=1,column=2).value = x.strftime('%x')
            sheet.cell(row=1,column=2).font = bold_font
            var = 2

            for ele in mydatalist:
                sheet.cell(row=var,column=2).value = 'ABSENT'
                sheet.cell(row=var,column=2).font = big_red_text
                var = var + 1
            workbook.save('sample.xlsx')
            
        
        if mydata in mydatalist:
            pathattendance = r'C:\Users\Dell\Desktop\fin\attendance'
            newpath = pathattendance + "\\" + str(mydata) 
            print(newpath)
            output = 'Authorised'
            mycolor = (0,255,0)
            var = 2
            isfolder = os.path.isdir(newpath)
            print(isfolder)

            if isfolder:
                os.chdir(newpath)
                imgname = str(mydata) + "_" + str(x.strftime('%d')) + "_" + str(x.strftime('%b')) + "_" + str(x.strftime('%Y')) + "_" +str(x.strftime('%H')) + "_" + str(x.strftime('%M')) + "_" + ".jpg"
                print(imgname)
                cv2.imwrite(imgname,img)
            else:
                os.mkdir(newpath)
                os.chdir(newpath)
                imgname = str(mydata) + "_" + str(x.strftime('%d')) + "_" + str(x.strftime('%b')) + "_" + str(x.strftime('%Y')) + "_" +str(x.strftime('%H')) + "_" + str(x.strftime('%M')) + "_" + ".jpg"
                print(imgname)
                cv2.imwrite(imgname,img)

            for ele in mydatalist:
                if sheet.cell(row=var,column=1).value == mydata :
                    sheet.cell(row=var,column=2).value = 'PRESENT ' + x.strftime('%X')
                    sheet.cell(row=var,column=2).font = big_green_text
                var = var + 1

            rootpath = r'C:\Users\Dell\Desktop\fin'
            os.chdir(rootpath)
            workbook.save('sample.xlsx')
        else:
            output = 'Not Authorised'
            mycolor = (0,0,255)

        pts = np.array([barcode.polygon],np.int32)
        pts = pts.reshape((-1,1,2))
        cv2.polylines(img,[pts],True,mycolor,5)
        pts2 = barcode.rect
        cv2.putText(img,output,(pts2[0],pts2[1]),cv2.FONT_HERSHEY_COMPLEX,0.9,mycolor,2) 

    cv2.imshow('result',img)

    if cv2.waitKey(1) & 0xFF == ord('e'):
        break